"""
Excel 匯入模組 - 讀取自訂預算表與 cwmoney 記帳紀錄
支援自動偵測最新 cwmoney 匯出檔、上傳檔案、Google Sheets 線上同步
"""
import os
import glob
import io
import openpyxl
import pandas as pd
from datetime import datetime, date as date_type
from typing import Optional, Union


# ─── 環境偵測 ─────────────────────────────────────────────
# Streamlit Cloud 上沒有本機資料夾，改用 repo 內的 data/ 目錄
_APP_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
_BUNDLED_DATA_DIR = os.path.join(_APP_DIR, "data")
_LOCAL_DATA_DIR = r"C:\Users\CHIUAL\Documents\Alan's budget"

# 智慧偵測：本機有資料夾就用本機，否則用 bundled data/
IS_CLOUD = not os.path.exists(_LOCAL_DATA_DIR)
DATA_DIR = _LOCAL_DATA_DIR if not IS_CLOUD else _BUNDLED_DATA_DIR

# 預算表路徑（依環境決定）
_BUDGET_FILENAME = "2026 New Monthly Budge.xlsx"
BUDGET_EXCEL_PATH = os.path.join(DATA_DIR, _BUDGET_FILENAME)

# 如果本機沒有但 bundled 有，也用 bundled
if not os.path.exists(BUDGET_EXCEL_PATH) and os.path.exists(os.path.join(_BUNDLED_DATA_DIR, _BUDGET_FILENAME)):
    BUDGET_EXCEL_PATH = os.path.join(_BUNDLED_DATA_DIR, _BUDGET_FILENAME)

# cwmoney 匯出檔的檔名前綴模式（用來自動偵測）
CWMONEY_FILE_PATTERN = "cwmoney_ex2_db_CSV_*.xlsx"


def find_cwmoney_files(data_dir: str = DATA_DIR) -> list:
    """
    掃描資料目錄，找出所有 cwmoney 匯出檔
    回傳按修改時間排序的檔案路徑列表（最新的在前）
    """
    pattern = os.path.join(data_dir, CWMONEY_FILE_PATTERN)
    files = glob.glob(pattern)
    # 按檔案修改時間排序（最新在前）
    files.sort(key=os.path.getmtime, reverse=True)
    return files


def get_latest_cwmoney_file(data_dir: str = DATA_DIR) -> Optional[str]:
    """
    自動取得最新的 cwmoney 匯出檔路徑
    回傳最新檔案的完整路徑，若無則回傳 None
    """
    files = find_cwmoney_files(data_dir)
    return files[0] if files else None


# 預設使用最新的 cwmoney 檔案（向後相容）
CWMONEY_EXCEL_PATH = get_latest_cwmoney_file() or ""


def load_budget_from_excel(filepath: str = BUDGET_EXCEL_PATH) -> dict:
    """
    從 Excel 預算表讀取預算結構

    回傳格式:
    {
        "categories": [
            {
                "project": "食",
                "main_category": "食品酒水",
                "sub_category": "伙食費",
                "expense_type": "固定支出",
                "budget": 11000
            },
            ...
        ],
        "sub_category_mapping": {
            "伙食費": ["早餐", "午餐", "晚餐", "牛奶", "食材費用", "水果", "水", "主食"],
            "零食費": ["飲料", "咖啡", "零食"],
            ...
        },
        "structure": {
            "食": {
                "食品酒水": {
                    "伙食費": {"budget": 11000, "type": "固定支出"},
                    "零食費": {"budget": 1000, "type": "變動支出"}
                }
            },
            ...
        }
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)

    # ── 讀取主預算表 ──
    ws = wb['2026 New Monthly budge_Alan']
    categories = []
    structure = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        project, main_cat, sub_cat, expense_type, budget_amount = row[:5]
        if not project or not main_cat or not sub_cat:
            continue

        budget_amount = budget_amount or 0

        categories.append({
            "project": str(project).strip(),
            "main_category": str(main_cat).strip(),
            "sub_category": str(sub_cat).strip(),
            "expense_type": str(expense_type).strip() if expense_type else "變動支出",
            "budget": float(budget_amount)
        })

        # 建立階層結構
        proj = str(project).strip()
        main = str(main_cat).strip()
        sub = str(sub_cat).strip()

        if proj not in structure:
            structure[proj] = {}
        if main not in structure[proj]:
            structure[proj][main] = {}
        structure[proj][main][sub] = {
            "budget": float(budget_amount),
            "type": str(expense_type).strip() if expense_type else "變動支出"
        }

    # ── 讀取子分類對應表 ──
    sub_category_mapping = {}
    if '對應預算子分類' in wb.sheetnames:
        ws_map = wb['對應預算子分類']
        for row in ws_map.iter_rows(min_row=2, max_row=ws_map.max_row, values_only=True):
            sub_cat, mapped_name = row[:2]
            if sub_cat and mapped_name:
                sub_cat = str(sub_cat).strip()
                mapped_name = str(mapped_name).strip()
                if sub_cat not in sub_category_mapping:
                    sub_category_mapping[sub_cat] = []
                sub_category_mapping[sub_cat].append(mapped_name)

    wb.close()

    return {
        "categories": categories,
        "sub_category_mapping": sub_category_mapping,
        "structure": structure
    }


def get_project_icons() -> dict:
    """專案層級的圖示對應"""
    return {
        "食": "🍜",
        "住": "🏠",
        "行": "🚗",
        "育": "📚",
        "大寶": "👶",
        "樂": "🎮",
        "衣": "👕",
        "儲蓄": "💰",
    }


def get_main_category_icons() -> dict:
    """主分類的圖示對應"""
    return {
        "食品酒水": "🍜",
        "金融保險": "🏦",
        "居家物業": "🏠",
        "行車交通": "🚗",
        "交流通訊": "📱",
        "進修學習": "📚",
        "人情往來": "💝",
        "醫療保健": "💊",
        "休閒娛樂": "🎮",
    }


def get_budget_summary_by_project(budget_data: dict) -> dict:
    """按專案層級彙總預算"""
    icons = get_project_icons()
    summary = {}
    for cat in budget_data["categories"]:
        proj = cat["project"]
        label = f"{icons.get(proj, '📌')} {proj}"
        summary[label] = summary.get(label, 0) + cat["budget"]
    return summary


def get_budget_summary_by_main_category(budget_data: dict) -> dict:
    """按主分類彙總預算"""
    icons = get_main_category_icons()
    summary = {}
    for cat in budget_data["categories"]:
        main = cat["main_category"]
        label = f"{icons.get(main, '📌')} {main}"
        summary[label] = summary.get(label, 0) + cat["budget"]
    return summary


def get_budget_summary_by_sub_category(budget_data: dict) -> dict:
    """按子分類列出預算"""
    summary = {}
    for cat in budget_data["categories"]:
        sub = cat["sub_category"]
        summary[sub] = cat["budget"]
    return summary


def get_all_main_categories(budget_data: dict) -> list:
    """取得所有主分類（不重複）"""
    seen = []
    for cat in budget_data["categories"]:
        if cat["main_category"] not in seen:
            seen.append(cat["main_category"])
    return seen


def get_sub_categories_for_main(budget_data: dict, main_category: str) -> list:
    """取得某主分類下的所有子分類"""
    subs = []
    for cat in budget_data["categories"]:
        if cat["main_category"] == main_category and cat["sub_category"] not in subs:
            subs.append(cat["sub_category"])
    return subs


def get_all_projects(budget_data: dict) -> list:
    """取得所有專案（不重複）"""
    seen = []
    for cat in budget_data["categories"]:
        if cat["project"] not in seen:
            seen.append(cat["project"])
    return seen


# ─── cwmoney 記帳紀錄匯入 ─────────────────────────────────

def load_cwmoney_records(filepath: str = CWMONEY_EXCEL_PATH,
                          year: Optional[int] = None,
                          month: Optional[int] = None) -> list:
    """
    從 cwmoney 匯出的 Excel 讀取記帳紀錄

    每筆紀錄格式:
    {
        "date": "2026-03-26",
        "type": "支出" | "收入" | "轉帳",
        "main_category": "食品酒水",
        "sub_category": "午餐",
        "account": "現金",
        "project": "食",
        "amount": 120,
        "note": "備註",
        "location": "地址",
        "invoice": "發票號碼"
    }
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb.worksheets[0]  # 取第一個 sheet

    records = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not row[0]:
            continue

        # 解析日期
        date_val = row[0]
        if isinstance(date_val, datetime):
            record_date = date_val.date()
        elif isinstance(date_val, str):
            try:
                record_date = datetime.strptime(date_val, "%Y-%m-%d").date()
            except ValueError:
                continue
        else:
            continue

        # 篩選年月
        if year and record_date.year != year:
            continue
        if month and record_date.month != month:
            continue

        record_type = str(row[1]).strip() if row[1] else ""
        main_cat = str(row[2]).strip() if row[2] else ""
        sub_cat = str(row[3]).strip() if row[3] else ""
        account = str(row[4]).strip() if row[4] else ""
        project = str(row[5]).strip() if row[5] else ""
        amount = float(row[6]) if row[6] else 0
        note = str(row[14]).strip() if row[14] and str(row[14]).strip() else ""
        location = str(row[11]).strip() if row[11] and str(row[11]).strip() else ""
        invoice = str(row[12]).strip() if row[12] and str(row[12]).strip() else ""

        records.append({
            "date": record_date.isoformat(),
            "type": record_type,
            "main_category": main_cat,
            "sub_category": sub_cat,
            "account": account,
            "project": project,
            "amount": amount,
            "note": note,
            "location": location,
            "invoice": invoice
        })

    wb.close()
    return records


# ─── cwmoney DataFrame 通用解析（上傳檔案 / Google Sheets 共用）─────

# cwmoney 匯出的標準欄位名稱（A~O 欄）
CWMONEY_COLUMNS = [
    "日期", "類別", "主分類", "子分類", "帳戶", "專案",
    "金額", "匯率", "小計", "建檔時間", "GPS", "地址",
    "發票號碼", "轉帳", "備註"
]


def parse_cwmoney_dataframe(df: pd.DataFrame,
                             year: Optional[int] = None,
                             month: Optional[int] = None) -> list:
    """
    從 pandas DataFrame 解析 cwmoney 紀錄（通用解析器）
    支援從上傳檔案或 Google Sheets 讀取的 DataFrame

    參數:
        df: cwmoney 格式的 DataFrame
        year: 篩選年份（None = 不篩選）
        month: 篩選月份（None = 不篩選）

    回傳: 同 load_cwmoney_records 的紀錄格式
    """
    if df is None or df.empty:
        return []

    # 嘗試標準化欄位名稱
    # 如果有 header，用欄位名稱對應；否則用位置索引
    col_map = {}
    cols = list(df.columns)

    # 嘗試用名稱對應
    name_to_idx = {
        "日期": 0, "類別": 1, "主分類": 2, "子分類": 3,
        "帳戶": 4, "專案": 5, "金額": 6, "地址": 11,
        "發票號碼": 12, "備註": 14
    }

    use_names = any(c in cols for c in ["日期", "類別", "主分類"])

    if use_names:
        for name, default_idx in name_to_idx.items():
            if name in cols:
                col_map[name] = name
            elif default_idx < len(cols):
                col_map[name] = cols[default_idx]
    else:
        # 純數字欄位，用索引
        for name, idx in name_to_idx.items():
            if idx < len(cols):
                col_map[name] = cols[idx]

    records = []
    for _, row in df.iterrows():
        # 解析日期
        date_val = row.get(col_map.get("日期", cols[0]))
        if pd.isna(date_val):
            continue

        if isinstance(date_val, (datetime, pd.Timestamp)):
            record_date = date_val.date() if hasattr(date_val, 'date') else date_val
        elif isinstance(date_val, date_type):
            record_date = date_val
        elif isinstance(date_val, str):
            try:
                record_date = datetime.strptime(date_val.strip(), "%Y-%m-%d").date()
            except ValueError:
                try:
                    record_date = datetime.strptime(date_val.strip(), "%Y/%m/%d").date()
                except ValueError:
                    continue
        else:
            continue

        # 篩選年月
        if year and record_date.year != year:
            continue
        if month and record_date.month != month:
            continue

        def safe_str(val):
            if pd.isna(val) or val is None:
                return ""
            return str(val).strip()

        def safe_float(val):
            if pd.isna(val) or val is None:
                return 0
            try:
                return float(val)
            except (ValueError, TypeError):
                return 0

        record_type = safe_str(row.get(col_map.get("類別", ""), ""))
        main_cat = safe_str(row.get(col_map.get("主分類", ""), ""))
        sub_cat = safe_str(row.get(col_map.get("子分類", ""), ""))
        account = safe_str(row.get(col_map.get("帳戶", ""), ""))
        project = safe_str(row.get(col_map.get("專案", ""), ""))
        amount = safe_float(row.get(col_map.get("金額", ""), 0))
        location = safe_str(row.get(col_map.get("地址", ""), ""))
        invoice = safe_str(row.get(col_map.get("發票號碼", ""), ""))
        note = safe_str(row.get(col_map.get("備註", ""), ""))

        records.append({
            "date": record_date.isoformat(),
            "type": record_type,
            "main_category": main_cat,
            "sub_category": sub_cat,
            "account": account,
            "project": project,
            "amount": amount,
            "note": note,
            "location": location,
            "invoice": invoice
        })

    return records


def load_cwmoney_from_uploaded_file(file_bytes: bytes,
                                     filename: str,
                                     year: Optional[int] = None,
                                     month: Optional[int] = None) -> list:
    """
    從上傳的檔案（xlsx 或 csv）解析 cwmoney 紀錄

    參數:
        file_bytes: 上傳檔案的 bytes 內容
        filename: 檔案名稱（用來判斷格式）
        year/month: 篩選年月
    """
    ext = os.path.splitext(filename)[1].lower()

    if ext == ".csv":
        df = pd.read_csv(io.BytesIO(file_bytes), encoding="utf-8-sig")
    elif ext in (".xlsx", ".xls"):
        df = pd.read_excel(io.BytesIO(file_bytes), engine="openpyxl")
    else:
        return []

    return parse_cwmoney_dataframe(df, year=year, month=month)


def load_budget_from_uploaded_file(file_bytes: bytes) -> dict:
    """
    從上傳的預算表 Excel 解析預算結構
    格式同 load_budget_from_excel 的回傳值
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)

    # 尋找主預算表 sheet
    budget_sheet = None
    for name in wb.sheetnames:
        if 'budge' in name.lower() or '預算' in name:
            budget_sheet = name
            break
    if not budget_sheet:
        budget_sheet = wb.sheetnames[0]

    ws = wb[budget_sheet]
    categories = []
    structure = {}

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        project, main_cat, sub_cat, expense_type, budget_amount = row[:5]
        if not project or not main_cat or not sub_cat:
            continue
        budget_amount = budget_amount or 0
        categories.append({
            "project": str(project).strip(),
            "main_category": str(main_cat).strip(),
            "sub_category": str(sub_cat).strip(),
            "expense_type": str(expense_type).strip() if expense_type else "變動支出",
            "budget": float(budget_amount)
        })
        proj = str(project).strip()
        main = str(main_cat).strip()
        sub = str(sub_cat).strip()
        if proj not in structure:
            structure[proj] = {}
        if main not in structure[proj]:
            structure[proj][main] = {}
        structure[proj][main][sub] = {
            "budget": float(budget_amount),
            "type": str(expense_type).strip() if expense_type else "變動支出"
        }

    sub_category_mapping = {}
    map_sheet = None
    for name in wb.sheetnames:
        if '對應' in name or 'mapping' in name.lower():
            map_sheet = name
            break
    if map_sheet:
        ws_map = wb[map_sheet]
        for row in ws_map.iter_rows(min_row=2, max_row=ws_map.max_row, values_only=True):
            sub_cat, mapped_name = row[:2]
            if sub_cat and mapped_name:
                sub_cat = str(sub_cat).strip()
                mapped_name = str(mapped_name).strip()
                if sub_cat not in sub_category_mapping:
                    sub_category_mapping[sub_cat] = []
                sub_category_mapping[sub_cat].append(mapped_name)

    wb.close()
    return {
        "categories": categories,
        "sub_category_mapping": sub_category_mapping,
        "structure": structure
    }


def load_cwmoney_from_google_sheets(sheet_url: str,
                                     year: Optional[int] = None,
                                     month: Optional[int] = None) -> list:
    """
    從 Google Sheets 公開連結讀取 cwmoney 紀錄

    支援的 URL 格式:
    - 完整分享連結: https://docs.google.com/spreadsheets/d/{ID}/edit...
    - 已發佈 CSV 連結: https://docs.google.com/spreadsheets/d/e/{ID}/pub?output=csv
    - 直接匯出連結: https://docs.google.com/spreadsheets/d/{ID}/export?format=csv

    使用方式:
    1. 在 Google Sheets 中開啟你的 cwmoney 資料
    2. 檔案 → 共用 → 發佈到網路 → 選擇 CSV 格式 → 發佈
    3. 將產生的連結貼到 App 中
    """
    csv_url = _convert_to_csv_url(sheet_url)

    try:
        df = pd.read_csv(csv_url, encoding="utf-8-sig")
        return parse_cwmoney_dataframe(df, year=year, month=month)
    except Exception as e:
        raise ConnectionError(f"無法讀取 Google Sheets: {e}")


def _convert_to_csv_url(url: str) -> str:
    """
    將各種 Google Sheets URL 格式轉換為 CSV 匯出 URL
    """
    import re
    url = url.strip()

    # 已經是 CSV 匯出格式
    if "output=csv" in url or "format=csv" in url:
        return url

    # 發佈格式: /d/e/XXXXX/pub
    match = re.search(r'/spreadsheets/d/e/([^/]+)/pub', url)
    if match:
        pub_id = match.group(1)
        return f"https://docs.google.com/spreadsheets/d/e/{pub_id}/pub?output=csv"

    # 編輯格式: /d/XXXXX/edit 或 /d/XXXXX
    match = re.search(r'/spreadsheets/d/([^/]+)', url)
    if match:
        sheet_id = match.group(1)
        # 嘗試從 URL 取得 gid（特定工作表）
        gid_match = re.search(r'gid=(\d+)', url)
        gid = gid_match.group(1) if gid_match else "0"
        return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"

    # 無法辨識，直接回傳（讓 pandas 嘗試讀取）
    return url


def get_cwmoney_monthly_summary(records: list, budget_data: dict) -> dict:
    """
    分析 cwmoney 紀錄，產生月度摘要（對照預算）

    回傳:
    {
        "total_income": 總收入,
        "total_expense": 總支出（不含儲蓄）,
        "total_savings": 儲蓄支出,
        "balance": 結餘,
        "expense_by_project": { 專案: 金額 },
        "expense_by_main_category": { 主分類: 金額 },
        "expense_by_sub_category": { 子分類: 金額 },
        "income_records": [...],
        "budget_vs_actual_by_sub": { 子分類: {budget, actual, diff, pct} },
        "budget_vs_actual_by_main": { 主分類: {budget, actual, diff, pct} },
        "budget_vs_actual_by_project": { 專案: {budget, actual, diff, pct} },
        "daily_expenses": { 日期: 金額 },
        "record_count": 總筆數,
        "expense_type_breakdown": { "固定支出": 金額, "變動支出": 金額 }
    }
    """
    icons_proj = get_project_icons()
    icons_main = get_main_category_icons()

    # 分類記錄
    expense_records = [r for r in records if r["type"] == "支出"]
    income_records = [r for r in records if r["type"] == "收入"]

    total_income = sum(r["amount"] for r in income_records)
    total_expense = sum(r["amount"] for r in expense_records)

    # 按專案統計
    expense_by_project = {}
    for r in expense_records:
        proj = r["project"]
        label = f"{icons_proj.get(proj, '📌')} {proj}"
        expense_by_project[label] = expense_by_project.get(label, 0) + r["amount"]

    # 按主分類統計
    expense_by_main = {}
    for r in expense_records:
        main = r["main_category"]
        label = f"{icons_main.get(main, '📌')} {main}"
        expense_by_main[label] = expense_by_main.get(label, 0) + r["amount"]

    # 按子分類統計
    expense_by_sub = {}
    for r in expense_records:
        sub = r["sub_category"]
        expense_by_sub[sub] = expense_by_sub.get(sub, 0) + r["amount"]

    # ── 預算 vs 實際（子分類層級）──
    # 建立 cwmoney 子分類 → 預算子分類的反向對應
    reverse_mapping = {}
    for budget_sub, mapped_items in budget_data.get("sub_category_mapping", {}).items():
        for item in mapped_items:
            reverse_mapping[item] = budget_sub

    # 將 cwmoney 的子分類金額歸入預算子分類
    expense_mapped_to_budget_sub = {}
    for r in expense_records:
        cwm_sub = r["sub_category"]
        # 先看是否有對應到預算子分類
        budget_sub = reverse_mapping.get(cwm_sub, cwm_sub)
        expense_mapped_to_budget_sub[budget_sub] = expense_mapped_to_budget_sub.get(budget_sub, 0) + r["amount"]

    budget_vs_actual_sub = {}
    for cat in budget_data["categories"]:
        sub = cat["sub_category"]
        budgeted = cat["budget"]
        actual = expense_mapped_to_budget_sub.get(sub, 0)
        budget_vs_actual_sub[sub] = {
            "budget": budgeted,
            "actual": actual,
            "diff": budgeted - actual,
            "pct": (actual / budgeted * 100) if budgeted > 0 else (100 if actual > 0 else 0),
            "project": cat["project"],
            "main_category": cat["main_category"],
            "expense_type": cat["expense_type"]
        }

    # ── 預算 vs 實際（主分類層級）──
    budget_by_main = {}
    actual_by_main = {}
    for cat in budget_data["categories"]:
        main = cat["main_category"]
        label = f"{icons_main.get(main, '📌')} {main}"
        budget_by_main[label] = budget_by_main.get(label, 0) + cat["budget"]
    for sub, data in budget_vs_actual_sub.items():
        main = data["main_category"]
        label = f"{icons_main.get(main, '📌')} {main}"
        actual_by_main[label] = actual_by_main.get(label, 0) + data["actual"]

    budget_vs_actual_main = {}
    for label in budget_by_main:
        budgeted = budget_by_main[label]
        actual = actual_by_main.get(label, 0)
        budget_vs_actual_main[label] = {
            "budget": budgeted,
            "actual": actual,
            "diff": budgeted - actual,
            "pct": (actual / budgeted * 100) if budgeted > 0 else (100 if actual > 0 else 0)
        }

    # ── 預算 vs 實際（專案層級）──
    budget_by_proj = {}
    actual_by_proj = {}
    for cat in budget_data["categories"]:
        proj = cat["project"]
        label = f"{icons_proj.get(proj, '📌')} {proj}"
        budget_by_proj[label] = budget_by_proj.get(label, 0) + cat["budget"]
    for r in expense_records:
        proj = r["project"]
        label = f"{icons_proj.get(proj, '📌')} {proj}"
        actual_by_proj[label] = actual_by_proj.get(label, 0) + r["amount"]

    budget_vs_actual_proj = {}
    for label in budget_by_proj:
        budgeted = budget_by_proj[label]
        actual = actual_by_proj.get(label, 0)
        budget_vs_actual_proj[label] = {
            "budget": budgeted,
            "actual": actual,
            "diff": budgeted - actual,
            "pct": (actual / budgeted * 100) if budgeted > 0 else (100 if actual > 0 else 0)
        }

    # 每日支出
    daily_expenses = {}
    for r in expense_records:
        d = r["date"]
        daily_expenses[d] = daily_expenses.get(d, 0) + r["amount"]

    # 固定/變動支出分類統計
    expense_type_breakdown = {}
    for sub, data in budget_vs_actual_sub.items():
        etype = data["expense_type"]
        expense_type_breakdown[etype] = expense_type_breakdown.get(etype, 0) + data["actual"]

    # 儲蓄相關
    total_savings_budget = sum(
        cat["budget"] for cat in budget_data["categories"]
        if cat["expense_type"] == "儲蓄支出"
    )
    budget_total_no_savings = sum(
        cat["budget"] for cat in budget_data["categories"]
        if cat["expense_type"] != "儲蓄支出"
    )
    budget_total = sum(cat["budget"] for cat in budget_data["categories"])

    # 實際儲蓄支出金額（從 cwmoney 記錄中歸類到儲蓄的部分）
    actual_savings_expense = expense_type_breakdown.get("儲蓄支出", 0)
    # 不含儲蓄的實際支出
    expense_no_savings = total_expense - actual_savings_expense

    return {
        "total_income": total_income,
        "total_expense": total_expense,
        "total_savings_budget": total_savings_budget,
        "actual_savings_expense": actual_savings_expense,
        "expense_no_savings": expense_no_savings,
        "balance": total_income - total_expense,
        "budget_total": budget_total,
        "budget_total_no_savings": budget_total_no_savings,
        "budget_remaining": budget_total_no_savings - expense_no_savings,
        "is_over_budget": expense_no_savings > budget_total_no_savings,
        "expense_by_project": expense_by_project,
        "expense_by_main_category": expense_by_main,
        "expense_by_sub_category": expense_by_sub,
        "income_records": income_records,
        "budget_vs_actual_by_sub": budget_vs_actual_sub,
        "budget_vs_actual_by_main": budget_vs_actual_main,
        "budget_vs_actual_by_project": budget_vs_actual_proj,
        "daily_expenses": daily_expenses,
        "record_count": len(records),
        "expense_count": len(expense_records),
        "income_count": len(income_records),
        "expense_type_breakdown": expense_type_breakdown
    }
